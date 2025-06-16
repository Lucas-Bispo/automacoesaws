import logging
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# --- MAPA DE LINKS FINAL ---
# Define como as abas devem se conectar umas com as outras
LINKING_CONFIG = {
    'Subnets': {'VpcId': 'VPCs'},
    'SecurityGroups': {'VpcId': 'VPCs', 'SourceDest': 'SecurityGroups'},
    'RouteTables': {'VpcId': 'VPCs', 'Target': 'InternetGateways'},
    'Security_Analysis': {'ID do Security Group': 'SecurityGroups'}
}

# --- CORES E FONTES PARA ANÁLISE E FORMATAÇÃO ---
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FONT = Font(color="9C0006", bold=True)
YELLOW_FONT = Font(color="9C6500", bold=True)
GREEN_FONT = Font(color="006100", bold=True)

def apply_final_formatting(input_path: str, output_path: str, sg_risk_map: dict):
    """
    Lê uma planilha analisada, aplica toda a formatação final (links, cores, layout)
    e salva o resultado no arquivo de saída final.
    """
    logging.info(f"Aplicando formatação final em: {os.path.basename(input_path)}")
    try:
        # Carrega o workbook a partir do arquivo de entrada
        workbook = load_workbook(input_path)
        
        # ETAPA 1: CRIAÇÃO DOS MAPAS DE IDS PARA OS LINKS
        id_maps = {sheet.title: {str(cell.value): cell.row for cell in sheet['A'] if cell.value is not None} for sheet in workbook if sheet.max_row > 1}

        # ETAPA 2: APLICAÇÃO DOS HYPERLINKS
        for source_sheet_name, column_mappings in LINKING_CONFIG.items():
            if source_sheet_name in workbook.sheetnames:
                sheet = workbook[source_sheet_name]
                try:
                    headers = [cell.value for cell in sheet[1]]
                    for source_col_name, target_sheet_name in column_mappings.items():
                        if source_col_name in headers and target_sheet_name in id_maps:
                            col_idx = headers.index(source_col_name) + 1
                            target_map = id_maps[target_sheet_name]
                            for row_num in range(2, sheet.max_row + 1):
                                cell = sheet.cell(row=row_num, column=col_idx)
                                if cell.value and isinstance(cell.value, str) and cell.value.strip() in target_map:
                                    link_id = cell.value.strip()
                                    target_row = target_map[link_id]
                                    link_location = f"#'{target_sheet_name}'!A{target_row}"
                                    cell.value = f'=HYPERLINK("{link_location}", "{cell.value}")'
                                    cell.style = "Hyperlink"
                except (ValueError, IndexError):
                     logging.warning(f"Não foi possível processar links para a aba '{source_sheet_name}'.")

        # ETAPA 3: COLORAÇÃO DAS ABAS
        # Colore as linhas da aba SecurityGroups com base no mapa de risco
        if 'SecurityGroups' in workbook.sheetnames and sg_risk_map:
            sheet = workbook['SecurityGroups']
            try:
                headers = [c.value for c in sheet[1]]
                # Assumindo que a coluna de ID do SG no relatório sumarizado é 'GroupId'
                sg_id_col_name = next((h for h in headers if 'GroupId' in h), None)
                if sg_id_col_name:
                    sg_id_col_idx = headers.index(sg_id_col_name) + 1
                    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                        sg_id_cell_value = row_cells[sg_id_col_idx - 1].value
                        # Extrai o ID se for um hyperlink
                        sg_id = str(sg_id_cell_value).split('"')[-2] if "HYPERLINK" in str(sg_id_cell_value) else str(sg_id_cell_value)
                        
                        risk_level = sg_risk_map.get(sg_id, "Seguro")
                        fill_style = {"Alto": RED_FILL, "Médio": YELLOW_FILL, "Seguro": GREEN_FILL}.get(risk_level)
                        
                        if fill_style:
                            for cell in row_cells:
                                cell.fill = fill_style
            except ValueError:
                logging.warning("Coluna 'GroupId' não encontrada para colorir linhas.")

        # ETAPA 4: LAYOUT GERAL (AJUSTE DE COLUNAS E ALINHAMENTO)
        for sheet in workbook:
            for col in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    if cell.value:
                        cell_text = str(cell.value)
                        # Se for hyperlink, pega só o texto visível para o cálculo
                        if "HYPERLINK" in cell_text:
                            try:
                                cell_text = cell_text.split('"')[-2]
                            except IndexError:
                                pass
                        
                        cell_max_line = max(len(line) for line in cell_text.split('\n'))
                        if cell_max_line > max_length:
                            max_length = cell_max_line
                
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 70)

        # Salva o workbook final no caminho de saída
        workbook.save(output_path)
        logging.info(f"Relatório final formatado e salvo com sucesso em: {os.path.basename(output_path)}")

    except FileNotFoundError:
        logging.error(f"Arquivo de entrada para formatação não encontrado: {input_path}"); raise
    except Exception as e:
        logging.error(f"Erro durante a formatação final: {e}", exc_info=True); raise