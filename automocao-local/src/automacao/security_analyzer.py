import pandas as pd
import logging
import json
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

HIGH_RISK_PORTS = {22, 3389, 3306, 5432, 1433, 27017}
ACCEPTABLE_PUBLIC_PORTS = {80, 443}
ANYWHERE_CIDR = '0.0.0.0/0'

def analyze_sgs(sg_dataframe: pd.DataFrame):
    """Analisa um DataFrame SUMARIZADO de SGs e retorna findings e mapa de risco."""
    logging.info("Analisando dados brutos de Security Groups para riscos...")
    findings, sg_risk_map = [], {}
    if sg_dataframe.empty:
        return pd.DataFrame([{"Risco": "Info", "ID do SG": "Nenhum SG encontrado."}]), {}
    for _, sg_row in sg_dataframe.iterrows():
        sg_id, highest_risk_level = sg_row.get('GroupId'), 0
        if not sg_id: continue
        for rule in sg_row.get('IpPermissions', []):
            if any(ip.get('CidrIp') == ANYWHERE_CIDR for ip in rule.get('IpRanges', [])):
                fp, tp, proto = rule.get('FromPort'), rule.get('ToPort'), str(rule.get('IpProtocol', '-1')).upper().replace('-1', 'All')
                if fp is not None and tp is not None:
                    for port in range(fp, tp + 1):
                        txt = f"Entrada {proto}:{port} de {ANYWHERE_CIDR}"
                        if port in HIGH_RISK_PORTS:
                            highest_risk_level = max(highest_risk_level, 2)
                            findings.append({"Risco": "Alto", "ID do Security Group": sg_id, "Nome do Grupo": sg_row.get('GroupName'), "Regra Problemática": txt, "Recomendação": "Acesso crítico exposto. RESTRINJA."})
                        elif port not in ACCEPTABLE_PUBLIC_PORTS:
                            highest_risk_level = max(highest_risk_level, 1)
                            findings.append({"Risco": "Médio", "ID do Security Group": sg_id, "Nome do Grupo": sg_row.get('GroupName'), "Regra Problemática": txt, "Recomendação": "Porta não-padrão exposta. Verifique."})
        sg_risk_map[sg_id] = "Alto" if highest_risk_level == 2 else "Médio" if highest_risk_level == 1 else "Seguro"
    findings_df = pd.DataFrame(findings) if findings else pd.DataFrame([{"Risco": "Parabéns!", "ID do Security Group": "Nenhum risco comum detectado."}])
    logging.info(f"Análise de segurança concluída. {len(findings)} riscos encontrados.")
    return findings_df, sg_risk_map

def analyze_and_update_report(input_path: str, output_path: str, json_path: str):
    logging.info(f"Analisando segurança a partir de: {os.path.basename(json_path)}")
    try:
        with open(json_path, 'r', encoding='utf-8') as f: raw_data = json.load(f)
        sg_df_raw = pd.DataFrame(raw_data.get('SecurityGroups', []))
        findings_df, sg_risk_map = analyze_sgs(sg_dataframe=sg_df_raw)
        
        workbook = load_workbook(input_path)
        if 'Security_Analysis' in workbook.sheetnames: del workbook['Security_Analysis']
        analysis_sheet = workbook.create_sheet('Security_Analysis', 0)
        for r in dataframe_to_rows(findings_df, index=False, header=True):
            analysis_sheet.append(r)
        
        workbook.save(output_path)
        logging.info(f"Relatório analisado salvo em: {os.path.basename(output_path)}")
        return sg_risk_map
    except FileNotFoundError:
        logging.error(f"Arquivo não encontrado: {input_path} ou {json_path}"); raise