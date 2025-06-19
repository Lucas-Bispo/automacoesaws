import pandas as pd  # Biblioteca para manipulação de dados tabulares (DataFrames)
import boto3  # Biblioteca oficial AWS para interagir com serviços AWS via API
import logging  # Biblioteca para registrar logs de eventos e erros
import io  # Biblioteca para manipulação de streams em memória (usada para criar arquivo Excel em memória)
import os  # Biblioteca para manipulação de arquivos e diretórios
from openpyxl import load_workbook  # Biblioteca para manipular arquivos Excel (.xlsx)
from openpyxl.styles import Alignment, PatternFill, Font  # Para formatar células Excel (alinhamento, cor, fonte)
from openpyxl.utils import get_column_letter  # Para converter número de coluna em letra (ex: 1 -> 'A')
from collections import defaultdict  # Estrutura de dados que cria dicionário com listas automaticamente
from ..models import VPC, SecurityGroup  # Importa classes que modelam VPC e Security Group
from ..utils import formatters  # Importa utilitários para formatar regras de segurança
from ..security_analyzer import analyze_sgs  # Importa função que analisa riscos dos Security Groups

class VPCReport:
    """Fábrica autônoma para criar o relatório completo de VPC em memória."""

    def __init__(self, regions_to_scan: list):
        # Recebe a lista de regiões AWS que serão escaneadas
        self.regions_to_scan = regions_to_scan
        
        # Inicializa lista que armazenará objetos VPC carregados da AWS
        self.vpcs: list[VPC] = []
        
        # Inicializa DataFrame vazio para armazenar os resultados da análise de segurança
        self.findings_df = pd.DataFrame()
        
        # Inicializa dicionário que mapeará ID do Security Group para seu nível de risco
        self.sg_risk_map = {}
        
        # Registra no log o início da fábrica com o número de regiões a escanear
        logging.info(f"Fábrica de Relatório VPC iniciada para {len(self.regions_to_scan)} região(ões).")

    def collect_data(self):
        """ETAPA 1: Coleta dados brutos da AWS, cria e interliga os objetos em memória."""
        logging.info("Iniciando coleta e construção do modelo de dados...")
        
        # Listas para armazenar dados brutos de Security Groups e VPCs de todas as regiões
        all_sgs_raw = []
        all_vpcs_raw = []
        
        # Cria uma sessão boto3 para interagir com AWS
        session = boto3.Session()

        # Para cada região definida para escanear
        for region in self.regions_to_scan:
            logging.info(f"Coletando dados da região: {region}...")
            try:
                # Cria cliente EC2 para a região atual
                client = session.client('ec2', region_name=region)
                
                # Coleta dados brutos das VPCs e Security Groups da região via API AWS
                vpcs_data = client.describe_vpcs().get('Vpcs', [])
                sgs_data = client.describe_security_groups().get('SecurityGroups', [])
                
                # Adiciona a informação da região em cada item coletado para referência futura
                for item in vpcs_data: item['Region'] = region
                for item in sgs_data: item['Region'] = region
                
                # Extende as listas globais com os dados coletados da região atual
                all_vpcs_raw.extend(vpcs_data)
                all_sgs_raw.extend(sgs_data)
            except Exception as e:
                # Caso falhe a coleta em alguma região, registra aviso e continua
                logging.warning(f"Falha ao coletar dados da região {region}: {e}")
        
        # Cria objetos VPC a partir dos dados brutos coletados
        vpcs_obj = [VPC(data) for data in all_vpcs_raw]
        
        # Cria objetos Security Group a partir dos dados brutos coletados
        sgs_obj = [SecurityGroup(data) for data in all_sgs_raw]
        
        # Agrupa os Security Groups por ID da VPC a que pertencem
        sgs_by_vpc = defaultdict(list)
        for sg in sgs_obj:
            sgs_by_vpc[sg.vpc_id].append(sg)
        
        # Para cada VPC, associa a lista de Security Groups correspondentes
        for vpc in vpcs_obj:
            vpc.security_groups = sgs_by_vpc.get(vpc.id, [])
        
        # Armazena a lista completa de VPCs com seus Security Groups no atributo da classe
        self.vpcs = vpcs_obj
        
        logging.info(f"Modelo de dados com {len(self.vpcs)} VPCs construído.")
        
        # Retorna self para permitir encadeamento de métodos (ex: factory.collect_data().analyze_security())
        return self

    def analyze_security(self):
        """ETAPA 2: Analisa os SGs coletados e armazena os resultados internamente."""
        logging.info("Analisando riscos de segurança dos objetos...")
        
        # Junta todos os Security Groups de todas as VPCs em uma lista única
        all_sgs_objects = [sg for vpc in self.vpcs for sg in vpc.security_groups]
        
        # Chama função externa para analisar os Security Groups e obter:
        # - DataFrame com achados da análise
        # - Mapeamento do risco de cada Security Group
        self.findings_df, self.sg_risk_map = analyze_sgs(all_sgs_objects)
        
        # Atualiza o atributo risk_level de cada Security Group com o resultado da análise
        for sg in all_sgs_objects:
            sg.risk_level = self.sg_risk_map.get(sg.id, "Seguro")
        
        logging.info("Análise de segurança concluída.")
        
        # Retorna self para encadeamento
        return self

    def generate_report(self, output_path: str):
        """ETAPA 3 e 4: Gera a planilha final, formatada, com links e a salva no disco."""
        logging.info("Gerando e formatando relatório final...")
        
        # Constrói DataFrames para cada aba do Excel a partir dos objetos em memória
        data_frames = self._build_dataframes()
        
        # Cria um buffer em memória para salvar o arquivo Excel temporariamente
        buffer = io.BytesIO()
        
        # Usa pandas ExcelWriter com engine openpyxl para escrever múltiplas abas
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Define ordem das abas no arquivo final
            sheet_order = ['Security_Analysis', 'VPCs', 'SecurityGroups']
            
            # Para cada aba, se existir dados, escreve no Excel
            for sheet_name in data_frames.keys():
                if sheet_name in sheet_order and not data_frames[sheet_name].empty:
                    data_frames[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Carrega o arquivo Excel do buffer para aplicar formatações avançadas
        workbook = load_workbook(buffer)
        workbook = self._apply_final_formatting(workbook)

        # Garante que o diretório de saída exista
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Salva o arquivo Excel final no disco
        workbook.save(output_path)
        
        logging.info(f"Relatório final gerado com sucesso em: {os.path.basename(output_path)}")

    def _build_dataframes(self):
        """Método privado para converter os objetos em DataFrames prontos para o Excel."""
        
        # Prepara lista de dicionários com dados das VPCs para o DataFrame
        vpcs_for_df = [{'VpcId': v.id, 'VPC Name': v.name, 'Region': v.region} for v in self.vpcs]
        
        # Prepara lista de dicionários com dados dos Security Groups, formatando regras para leitura
        sgs_for_df = [
            {
                'GroupId': sg.id,
                'GroupName': sg.name,
                'VpcId': sg.vpc_id,
                'Region': sg.region,
                'Inbound Rules': formatters.format_rules(sg.raw_rules.get('IpPermissions', [])),
                'Outbound Rules': formatters.format_rules(sg.raw_rules.get('IpPermissionsEgress', []))
            }
            for v in self.vpcs for sg in v.security_groups
        ]
        
        # Retorna um dicionário com os DataFrames para cada aba do Excel
        return {
            'VPCs': pd.DataFrame(vpcs_for_df),
            'SecurityGroups': pd.DataFrame(sgs_for_df),
            'Security_Analysis': self.findings_df
        }

    def _apply_final_formatting(self, workbook):
        """Método privado para aplicar cores de linha, hyperlinks e layout no Excel."""
        
        # Configurações para criar links entre abas (não implementado totalmente aqui)
        LINKING_CONFIG = {
            'SecurityGroups': {'VpcId': 'VPCs'},
            'Security_Analysis': {'ID do Security Group': 'SecurityGroups'}
        }
        
        # Define cores para os níveis de risco: alto (vermelho), médio (amarelo), seguro (verde)
        RED_FILL = PatternFill(start_color='FFC7CE', fill_type='solid')
        YELLOW_FILL = PatternFill(start_color='FFEB9C', fill_type='solid')
        GREEN_FILL = PatternFill(start_color='C6EFCE', fill_type='solid')
        
        # Cria mapas de IDs para cada aba, associando valor da célula da coluna A para número da linha
        id_maps = {
            sheet.title: {str(cell.value): cell.row for cell in sheet['A'] if cell.value is not None}
            for sheet in workbook
        }

        # Aplica cores nas linhas da aba SecurityGroups conforme o nível de risco de cada SG
        if 'SecurityGroups' in workbook.sheetnames and self.sg_risk_map:
            sheet = workbook['SecurityGroups']
            try:
                # Encontra índice da coluna GroupId na primeira linha (cabeçalho)
                sg_id_col_idx = [c.value for c in sheet[1]].index('GroupId') + 1
                
                # Para cada linha (exceto cabeçalho), colore a linha conforme risco
                for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    sg_id = str(row_cells[sg_id_col_idx - 1].value)
                    risk_level = self.sg_risk_map.get(sg_id, "Seguro")
                    fill = {"Alto": RED_FILL, "Médio": YELLOW_FILL, "Seguro": GREEN_FILL}.get(risk_level)
                    if fill:
                        for cell in row_cells:
                            cell.fill = fill
            except ValueError:
                logging.warning("Coluna 'GroupId' não encontrada para colorir linhas.")
        
        # Ajusta largura das colunas e alinhamento do texto para todas as abas
        for sheet in workbook:
            for col in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Define alinhamento para texto quebrado, topo e esquerda
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    
                    # Calcula comprimento máximo do conteúdo da célula para ajustar largura da coluna
                    if cell.value:
                        max_length = max(max_length, max(len(line) for line in str(cell.value).split('\n')))
                
                # Ajusta largura da coluna com limite máximo para não ficar muito larga
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 70)

        # Retorna o workbook formatado para salvar
        return workbook
